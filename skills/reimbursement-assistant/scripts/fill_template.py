import argparse
from pathlib import Path

from lib_reimbursement import load_yaml_or_json, nested_get, read_json


def set_cell(ws, address, value):
    if value is None:
        value = ""
    ws[address] = value


def fill(input_json, template, map_path, output):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("fill_template.py requires openpyxl.") from exc

    data = read_json(input_json)
    template_map = load_yaml_or_json(map_path)
    wb = load_workbook(template)

    records = [r for r in data.get("records", []) if r.get("included")]
    for sheet_name, sheet_map in (template_map.get("sheets") or {}).items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for dotted, address in (sheet_map.get("fields") or {}).items():
            set_cell(ws, address, nested_get(data, dotted))

        table = sheet_map.get("expense_table") or {}
        start_row = int(table.get("start_row") or 0)
        max_rows = int(table.get("max_rows") or len(records))
        columns = table.get("columns") or {}
        if start_row and columns:
            for idx, record in enumerate(records[:max_rows]):
                row = start_row + idx
                for field, col in columns.items():
                    value = nested_get(record, field) if "." in field else record.get(field)
                    set_cell(ws, f"{col}{row}", value)

    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="Fill a company reimbursement Excel template.")
    parser.add_argument("--input", required=True, help="Confirmed/classified claim JSON.")
    parser.add_argument("--template", required=True, help="Excel template path.")
    parser.add_argument("--map", required=True, help="Template map JSON or YAML.")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    args = parser.parse_args()
    fill(args.input, args.template, args.map, args.output)


if __name__ == "__main__":
    main()

